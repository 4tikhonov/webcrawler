#!/usr/bin/python

import pandas as pd
import datetime
import time
import os
import re
import json
import simplejson
import openpyxl
from openpyxl.cell import get_column_letter
import random
import config

filenames = []
for filename in os.listdir(path):
    if re.search(r'bigdata', filename):
	filenames.append(filename)

def create_excel_dataset(fullpath, data):
    print "Creating %s" % fullpath
    wb = openpyxl.Workbook(encoding='utf-8')
    #ws = wb.get_active_sheet()
    ws = wb.active()
    ws.title = "Data"

    i = 1
    for i, item in result.iterrows():
        j = 1
        for value in item:
            c = ws.cell(row=i, column=j)
	    c.font = cell.font.copy(bold=True, italic=True)
            c.value = str(value).encode("utf-8")
            j+=1
        i+=1

#    wb.save(fullpath)
    return fullpath

def readlinks(path):
    xls_file = pd.ExcelFile(linksfile)
    sheets = xls_file.sheet_names
    for sheetname in sheets:
        df = xls_file.parse(sheetname, index_col=None)
    
    d = [
    dict([
    (colname, row[i])
    for i,colname in enumerate(df)
    ])
    for row in df.values
    ] 
    
    lexicon = {}
    for item in d:
        lexicon[item['link']] = item
    return lexicon
    
def datasets(path, filenames):
    frames = []
    for filename in filenames:
	print "Reading %s " % filename
        xl = pd.ExcelFile("%s/%s" % (path, filename))
        sheets = xl.sheet_names
        for sheetname in sheets:
            df = xl.parse(sheetname, index_col=None)
            df = df.fillna('')
            df.columns = ['asin', 'title','url','country','price','seller','currency','date','uid']
            info = df.head()
            data = df
        frames.append(data)
    return frames

# <codecell>

links = readlinks('links')

# <codecell>

def create_dataset(fullpath, data):
    wb = openpyxl.Workbook(encoding='utf-8')
    ws = wb.get_active_sheet()
    ws.title = "Data"

    asin = random.choice(properties.keys())      
    # Header
    rowid = 0
    if asin:
        cellid = 0
        for field in sorted(properties[asin]):            
            value = properties[asin][field]
            if field == 'yprices':
                prices = value
                for month in sorted(prices):                    
                    c = ws.cell(row=rowid, column=cellid)
                    c.value = str(month).encode("utf-8")                                        
                    ws.cell(row=rowid, column=cellid).style.font.bold = True
                    cellid+=1
            else:
                c = ws.cell(row=rowid, column=cellid)
                ws.cell(row=rowid, column=cellid).style.font.bold = True
                c.value = str(field).encode("utf-8")
            cellid+=1

    # Data
    rowid = 1
    for asin in properties:
        cellid = 0
        for field in sorted(properties[asin]):            
            value = properties[asin][field]
            if field == 'yprices':
                prices = value
                for month in sorted(prices):                    
                    c = ws.cell(row=rowid, column=cellid)                                                            
                    c.value = str(prices[month]).encode("utf-8")                    
                    cellid+=1
            else:                
                if cellid:
                    if ws.column_dimensions[get_column_letter(cellid+1)].width < len(str(value)):                    
                        ws.column_dimensions[get_column_letter(cellid+1)].width = len(str(value))                   
                else:
                    ws.column_dimensions[get_column_letter(cellid+1)].width = 10
                c = ws.cell(row=rowid, column=cellid)
                c.value = str(value).encode("utf-8")
            cellid+=1
        rowid+=1
    

    wb.save(fullpath)
    return fullpath

def getmonth():
    now = datetime.datetime.now()
    monname = time.strftime("%B")
    year = now.year
    if int(now.month) < 10:
        month = "0%s" % now.month
    else:
        month = "%s" % now.month
    monthpat = "%s-%s" % (str(year), str(month))
    return (monthpat, monname, year)

# <codecell>

frames = datasets(path, filenames)
result = pd.concat(frames)
(cat1,cat2,cat3,cat4,product,months01,months02,months03) = ([],[],[],[],[],[],[],[])

count = 0
properties = {}
monthslist = ["2016-01", "2016-02", "2016-03","2016-04","2016-05"]
(monthpat, monname, year) = getmonth()
monthslist.append(monthpat)

for i, item in result.iterrows():
    link = item['url']
    link = re.sub(r"&page\=\d+","",link)
    date = item['date']
    asin = item['asin']
    
    if asin not in properties:
        # New product
        thismonthprice = item['price']
        directurl = "%s/%s" % (producturl, asin)
        productitem = {}
        prices = {}
        totalcount = 0
        if asin:
            productitem['url'] = link        
            productitem['producturl'] = directurl                    
            month = re.match(r"(\d+)\-(\d+)", date)
            if month.group:
                yearmonth = month.group(0)
                count+=1
                prices[yearmonth] = item['price']
            for month in monthslist:
                if month not in prices:
                    prices[month] = ' '
                else:
                    totalcount+=1
            
            productitem['yprices'] = prices
            productitem['country'] = item['country']
            productitem['asin'] = item['asin']
            productitem['asin title'] = item['title']
            productitem['currency'] = item['currency']
            productitem['seller'] = item['seller']
            productitem['observations'] = totalcount
        
            try:
                productitem['category1'] = links[link]['category1'] 
            except:
                productitem['category1'] = 'None'
            
            try:
                productitem['category2'] = links[link]['category2'] 
            except:
                productitem['category2'] = 'None'

            try:
                productitem['category3'] = links[link]['category3'] 
            except:
                productitem['category3'] = 'None'

            try:
                productitem['category4'] = links[link]['category4'] 
            except:
                productitem['category4'] = 'None'
            
        properties[asin] = productitem
    else:
        # Product already exists
        month = re.match(r"(\d+)\-(\d+)", date)
        if month.group:
            yearmonth = month.group(0)
            count+=1
            properties[asin]['yprices'][yearmonth] = item['price']            
            properties[asin]['observations']+=1        

asin = random.choice(properties.keys())    
# B013B7O946
print json.dumps(properties[asin])
if asin:
    cellid = 0
    for field in sorted(properties[asin]):
        cellid+=1
        value = properties[asin][field]
        if field == 'yprices':
            prices = value
            for month in sorted(prices):
                print "%s %s" % (month, prices[month])
                cellid+=1
        else:
            print "%s %s" % (field, cellid)
    
create_dataset("%s/%s" % (finalpath, "finaldataset.xlsx"), properties)

# <codecell>


